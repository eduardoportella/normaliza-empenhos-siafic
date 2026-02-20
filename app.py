import io
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from dateutil import parser as dateparser
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import zipfile
import xml.etree.ElementTree as ET

def _strip_ns(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag

def read_weird_ooxml_xlsx(uploaded_file) -> pd.DataFrame:
    """
    Lê XLSX com namespace purl.oclc.org/ooxml (openpyxl costuma retornar 0 sheets).
    Retorna DataFrame cru (sem header), parecido com pd.read_excel(header=None).
    """
    # UploadedFile do streamlit é file-like; precisamos de bytes
    data = uploaded_file.getvalue()
    zf = zipfile.ZipFile(io.BytesIO(data))

    # sharedStrings
    shared = []
    if "xl/sharedStrings.xml" in zf.namelist():
        ss_xml = zf.read("xl/sharedStrings.xml")
        root = ET.fromstring(ss_xml)
        for si in root.findall(".//{*}si"):
            # concatena todos os <t> dentro de <si>
            texts = []
            for t in si.findall(".//{*}t"):
                if t.text:
                    texts.append(t.text)
            shared.append("".join(texts))

    # workbook -> pega o primeiro sheet r:id
    wb_xml = zf.read("xl/workbook.xml")
    wb_root = ET.fromstring(wb_xml)

    sheets = wb_root.findall(".//{*}sheet")
    if not sheets:
        raise ValueError("Não encontrei nenhuma aba no workbook.xml (arquivo inválido).")

    first_sheet = sheets[0]
    rid = first_sheet.attrib.get("{http://purl.oclc.org/ooxml/officeDocument/relationships}id") or first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
    if not rid:
        # fallback: assume sheet1.xml
        sheet_path = "xl/worksheets/sheet1.xml"
    else:
        # resolve rId -> worksheet file
        rels_xml = zf.read("xl/_rels/workbook.xml.rels")
        rels_root = ET.fromstring(rels_xml)
        target = None
        for rel in rels_root.findall(".//{*}Relationship"):
            if rel.attrib.get("Id") == rid:
                target = rel.attrib.get("Target")
                break
        sheet_path = "xl/" + target if target else "xl/worksheets/sheet1.xml"

    if sheet_path not in zf.namelist():
        raise ValueError(f"Worksheet XML não encontrado: {sheet_path}")

    sh_xml = zf.read(sheet_path)
    sh_root = ET.fromstring(sh_xml)

    # Lê células
    cells = {}
    max_r = 0
    max_c = 0

    def col_to_idx(col_letters: str) -> int:
        idx = 0
        for ch in col_letters:
            idx = idx * 26 + (ord(ch.upper()) - 64)
        return idx  # 1-based

    for c in sh_root.findall(".//{*}c"):
        ref = c.attrib.get("r")  # ex: A1
        if not ref:
            continue

        m = re.match(r"([A-Z]+)(\d+)", ref, re.I)
        if not m:
            continue

        col_letters, row_num = m.group(1), int(m.group(2))
        col_num = col_to_idx(col_letters)

        max_r = max(max_r, row_num)
        max_c = max(max_c, col_num)

        t = c.attrib.get("t")  # s, inlineStr, etc
        v_el = c.find("{*}v")
        is_el = c.find("{*}is")

        val = None
        if t == "s" and v_el is not None:
            # shared string
            try:
                val = shared[int(v_el.text)]
            except Exception:
                val = v_el.text
        elif t == "inlineStr" and is_el is not None:
            # inline string
            texts = []
            for tnode in is_el.findall(".//{*}t"):
                if tnode.text:
                    texts.append(tnode.text)
            val = "".join(texts)
        else:
            # num / str direto
            if v_el is not None:
                val = v_el.text

        cells[(row_num, col_num)] = val

    # monta matriz
    data_rows = []
    for r in range(1, max_r + 1):
        row = []
        for cidx in range(1, max_c + 1):
            row.append(cells.get((r, cidx), None))
        data_rows.append(row)

    return pd.DataFrame(data_rows)


# -----------------------------
# Config / helpers
# -----------------------------
FONT_ESTADUAL = 500
FONT_FEDERAL = 552

MONEY_FMT = '"R$" #,##0.00'
DATE_FMT = 'dd/mm/yyyy'

def clean_siafic_table(df_raw: pd.DataFrame, start_row: int = 3) -> pd.DataFrame:
    """
    - Remove as 3 primeiras linhas (start_row=3 => começa na linha 4 do Excel)
    - Assume que a linha 4 contém o cabeçalho
    - Remove a linha 'TOTAL' e tudo abaixo
    - Remove linhas de observação/rodapé (ex: 'SIAFIC PR - SISTEMA INTEGRADO...')
    """
    if df_raw is None or df_raw.empty:
        return df_raw

    # Garante que temos índice 0..n
    df_raw = df_raw.reset_index(drop=True)

    # Se veio com header já aplicado (colunas "bonitinhas"), a gente só corta topo/rodapé por linhas
    # Mas para padronizar, vamos tratar como "cru" quando o pandas leu com header=None.
    # Se as colunas são 0..n-1 (int), é bem provável que está cru.
    is_raw = all(isinstance(c, int) for c in df_raw.columns)

    if is_raw:
        # corta as 3 primeiras linhas e usa a próxima como cabeçalho
        if len(df_raw) <= start_row:
            return pd.DataFrame()

        header = df_raw.iloc[start_row].astype(str).str.strip().tolist()
        df = df_raw.iloc[start_row + 1:].copy()
        df.columns = header
        df = df.reset_index(drop=True)
    else:
        # já veio com cabeçalho: corta só as 3 primeiras linhas
        df = df_raw.iloc[start_row:].copy().reset_index(drop=True)

    # Remove colunas totalmente vazias
    df = df.dropna(axis=1, how="all")

    # --- Corta rodapé: primeira linha que indique "TOTAL" ou "SIAFIC PR - ..."
    # Procura em todas as colunas (porque às vezes o texto aparece deslocado)
    marker_patterns = [
        r"^\s*total\s*$",
        r"^\s*total\s+geral\s*$",
        r"siafic\s*pr\s*-\s*sistema\s+integrado",
    ]

    stop_idx = None
    # transforma em strings para varrer com segurança
    df_str = df.astype(str)

    for i in range(len(df_str)):
        row_text = " | ".join(df_str.iloc[i].tolist()).lower()
        if any(re.search(p, row_text, flags=re.IGNORECASE) for p in marker_patterns):
            stop_idx = i
            break

    if stop_idx is not None:
        df = df.iloc[:stop_idx].copy()

    # Remove linhas totalmente vazias (depois do corte)
    df = df.dropna(axis=0, how="all").reset_index(drop=True)

    return df

def clean_digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def format_cnpj_cpf(s: str) -> str:
    d = clean_digits(s)
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    if len(d) == 11:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    return d

def parse_date_any(v) -> Optional[datetime]:
    if pd.isna(v) or v is None or str(v).strip() == "":
        return None
    # Excel serial
    if isinstance(v, (int, float)) and v > 59 and v < 60000:
        # Excel base 1899-12-30
        return datetime(1899, 12, 30) + pd.to_timedelta(int(v), unit="D")
    s = str(v).strip()
    # dd/mm/yyyy
    try:
        dt = dateparser.parse(s, dayfirst=True, fuzzy=True)
        return dt
    except Exception:
        return None

def to_number_br(v) -> float:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return 0.0
    # remove currency and spaces
    s = re.sub(r"[^\d,.-]", "", s)
    # thousands '.' and decimal ','
    if s.count(",") == 1 and s.count(".") >= 1:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def normalize_header(h: str) -> str:
    h = (h or "").strip().lower()
    h = re.sub(r"\s+", " ", h)
    return h


def guess_columns(df: pd.DataFrame) -> Dict[str, str]:
    """
    Tenta achar colunas do SIAFIC pelo nome. Se não achar, cai no fallback por índice (igual Office Script).
    """
    cols = {c: normalize_header(c) for c in df.columns}
    # candidatos por “intenção”
    def find_any(keys: List[str]) -> Optional[str]:
        for c, lc in cols.items():
            for k in keys:
                if k in lc:
                    return c
        return None
    
    def find_fonte_col() -> Optional[str]:
      # 1) preferir header exatamente "fonte"
      for c, lc in cols.items():
         if lc == "fonte":
               return c

      # 2) aceitar variações que contenham "fonte" mas NÃO contenham "identificador" ou "exercício"
      candidates = []
      for c, lc in cols.items():
         if "fonte" in lc and "identificador" not in lc and "exerc" not in lc:
               candidates.append(c)

      # se tiver mais de uma, devolve a mais curta (normalmente é a coluna "Fonte")
      if candidates:
         return sorted(candidates, key=lambda x: len(str(x)))[0]
      return None 

    mapping = {
        "protocolo": find_any(["processo do empenho", "protocolo", "processo", "nº processo"]),
        "fonte": find_fonte_col(),
        "empenho": find_any(["nota de empenho"]),
        "data": find_any(["data da emissão do empenho"]),
        "valor": find_any(["despesas empenhadas"]),
        "cnpj": find_any(["cnpj", "cpf", "código credor", "cod credor"]),
        "credor": find_any(["descrição credor"]),
    }

    # fallback por índice (se a planilha vier “crua” tipo export)
    # indices do seu script: fonte=17, empenho=26, data=27, protocolo=30, cnpj=31, credor=32, valor=37
    # Aqui só funciona se o df vier com as mesmas colunas (muito comum no mesmo relatório)
    if any(v is None for v in mapping.values()):
        idx_map = {
            "fonte": 17,
            "empenho": 26,
            "data": 27,
            "protocolo": 30,
            "cnpj": 31,
            "credor": 32,
            "valor": 37,
        }
        for k, idx in idx_map.items():
            if mapping.get(k) is None and len(df.columns) > idx:
                mapping[k] = df.columns[idx]

    return mapping

def validate_mapping(mapping: Dict[str, str], df: pd.DataFrame) -> Tuple[bool, List[str]]:
    required = ["protocolo", "fonte", "empenho", "data", "valor"]
    missing = [k for k in required if not mapping.get(k) or mapping[k] not in df.columns]
    return (len(missing) == 0, missing)


# -----------------------------
# Core processing
# -----------------------------
def build_consolidated(
    df_siafic: pd.DataFrame,
    protocolos_input: pd.DataFrame,
    mapping: Dict[str, str]
) -> Tuple[pd.DataFrame, int, int]:
    """
    Retorna:
      - df_output "wide" já no layout final
      - maxEstadual
      - maxFederal
    """
    # normaliza entrada do usuário
    p = protocolos_input.copy()
    p.columns = [c.strip() for c in p.columns]
    # aceita nomes variados
    col_prot = next((c for c in p.columns if normalize_header(c) in ["protocolo", "processo", "processo do empenho"]), None)
    col_gms = next((c for c in p.columns if "gms" in normalize_header(c)), None)
    col_contrato = next((c for c in p.columns if "contrato" in normalize_header(c)), None)
    col_valcontr = next((c for c in p.columns if "valor" in normalize_header(c) and "contrato" in normalize_header(c)), None)

    if not col_prot or not col_gms or not col_contrato or not col_valcontr:
        raise ValueError("A tabela de entrada precisa ter colunas: Protocolo, GMS, Contrato e Valor do Contrato.")

    p[col_prot] = p[col_prot].astype(str).str.strip()
    p = p[p[col_prot] != ""]
    if p.empty:
        raise ValueError("Informe pelo menos 1 protocolo na tabela de entrada.")

    prot_map = {}
    for _, r in p.iterrows():
        prot = str(r[col_prot]).strip()
        prot_map[prot] = {
            "gms": str(r[col_gms]).strip() if not pd.isna(r[col_gms]) else "",
            "contrato": str(r[col_contrato]).strip() if not pd.isna(r[col_contrato]) else "",
            "valorContrato": to_number_br(r[col_valcontr]) if col_valcontr and col_valcontr in p.columns else None,
        }

    # filtra SIAFIC só para protocolos informados (reduz MUITO volume)
    df = df_siafic.copy()
    df["__prot__"] = df[mapping["protocolo"]].astype(str).str.strip()
    df = df[df["__prot__"].isin(set(prot_map.keys()))].copy()

    # estrutura por protocolo
    grouped: Dict[str, Dict] = {}
    for prot in prot_map.keys():
        grouped[prot] = {
            "credor": "Protocolo não encontrado no relatório do SIAFIC",
            "cnpj": "",
            "contrato": prot_map[prot]["contrato"],
            "gms": prot_map[prot]["gms"],
            "valorContrato": prot_map[prot]["valorContrato"],
            "estadual": [],
            "federal": [],
        }

    # percorre linhas relevantes
    for _, r in df.iterrows():
        prot = str(r["__prot__"]).strip()
        fonte = int(to_number_br(r[mapping["fonte"]]))
        emp = str(r[mapping["empenho"]] if mapping["empenho"] in r else "").strip()
        dt = parse_date_any(r[mapping["data"]])
        valor = float(to_number_br(r[mapping["valor"]]))

        if mapping.get("credor") and mapping["credor"] in df.columns:
            credor = str(r[mapping["credor"]] or "").upper().strip()
            if credor:
                grouped[prot]["credor"] = re.sub(r"\s+", " ", credor)

        if mapping.get("cnpj") and mapping["cnpj"] in df.columns:
            cnpj = str(r[mapping["cnpj"]] or "").strip()
            if cnpj:
                grouped[prot]["cnpj"] = clean_digits(cnpj)

        item = {"emp": emp, "data": dt, "valor": valor}


        print("\033[91m", fonte, "\033[0m", sep="")

        if fonte == FONT_ESTADUAL:
            grouped[prot]["estadual"].append(item)
        elif fonte == FONT_FEDERAL:
            grouped[prot]["federal"].append(item)

    # ordena por data
    for prot in grouped:
        grouped[prot]["estadual"].sort(key=lambda x: (x["data"] or datetime(1900, 1, 1)))
        grouped[prot]["federal"].sort(key=lambda x: (x["data"] or datetime(1900, 1, 1)))

    max_est = max((len(grouped[k]["estadual"]) for k in grouped), default=0)
    max_fed = max((len(grouped[k]["federal"]) for k in grouped), default=0)


    # monta dataframe wide
    rows = []
    for prot in sorted(grouped.keys()):
        g = grouped[prot]

        base = {
            "Protocolo": prot,
            "Credor": g["credor"],
            "CNPJ / CPF": format_cnpj_cpf(g["cnpj"]),
            "Contrato": g["contrato"],
            "GMS": g["gms"],
            "R$ Contrato": g["valorContrato"] if g["valorContrato"] is not None else None,
        }

        # estadual cols
        for i in range(max_est):
            emp = g["estadual"][i]["emp"] if i < len(g["estadual"]) else ""
            dt = g["estadual"][i]["data"] if i < len(g["estadual"]) else None
            val = g["estadual"][i]["valor"] if i < len(g["estadual"]) else None
            base[f"EST_EMP{i+1}"] = emp
            base[f"EST_DATA{i+1}"] = dt
            base[f"EST_VAL{i+1}"] = val

        # federal cols
        for i in range(max_fed):
            emp = g["federal"][i]["emp"] if i < len(g["federal"]) else ""
            dt = g["federal"][i]["data"] if i < len(g["federal"]) else None
            val = g["federal"][i]["valor"] if i < len(g["federal"]) else None
            base[f"FED_EMP{i+1}"] = emp
            base[f"FED_DATA{i+1}"] = dt
            base[f"FED_VAL{i+1}"] = val

        # totais
        total_est = sum([x["valor"] for x in g["estadual"]]) if g["estadual"] else 0.0
        total_fed = sum([x["valor"] for x in g["federal"]]) if g["federal"] else 0.0
        total_emp = total_est + total_fed
        base["TOTAL EST"] = total_est
        base["TOTAL FED"] = total_fed
        base["TOTAL EMP"] = total_emp

        if g["valorContrato"] is not None:
            base["Saldo"] = float(g["valorContrato"]) - total_emp
        else:
            base["Saldo"] = None

        rows.append(base)

    out = pd.DataFrame(rows)
    return out, max_est, max_fed


# -----------------------------
# Excel writer with formatting
# -----------------------------
def write_excel_formatted(df_out: pd.DataFrame, max_est: int, max_fed: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "REL CONSOLIDADO"

    # header rows (2 linhas)
    header1 = ["Protocolo", "Credor", "CNPJ / CPF", "Contrato", "GMS", "R$ Contrato"]
    header2 = ["", "", "", "", "", ""]

    for i in range(max_est):
        header1 += ["", "", ""]
        header2 += [f"EMP{i+1}", "DATA", "VALOR"]

    for i in range(max_fed):
        header1 += ["", "", ""]
        header2 += [f"EMP{i+1}", "DATA", "VALOR"]

    header1 += ["TOTAL EST", "TOTAL FED", "TOTAL EMP", "Saldo"]
    header2 += ["", "", "", ""]

    ws.append(header1)
    ws.append(header2)

    total_cols = len(header1)

    # merge ESTADUAL / FEDERAL titles on row1
    def col_letter(idx0: int) -> str:
        return get_column_letter(idx0 + 1)

    if max_est > 0:
        start = 6  # col index 0-based: G = 6? (A=0, F=5, G=6)
        end = 6 + max_est * 3 - 1
        ws.merge_cells(f"{col_letter(start)}1:{col_letter(end)}1")
        ws[f"{col_letter(start)}1"].value = "ESTADUAL"

    if max_fed > 0:
        start = 6 + max_est * 3
        end = start + max_fed * 3 - 1
        ws.merge_cells(f"{col_letter(start)}1:{col_letter(end)}1")
        ws[f"{col_letter(start)}1"].value = "FEDERAL"

    # write data rows
    for _, r in df_out.iterrows():
        row = []
        row += [
            r.get("Protocolo", ""),
            r.get("Credor", ""),
            r.get("CNPJ / CPF", ""),
            r.get("Contrato", ""),
            r.get("GMS", ""),
            r.get("R$ Contrato", None),
        ]

        # estaduais
        for i in range(max_est):
            row += [
                r.get(f"EST_EMP{i+1}", ""),
                r.get(f"EST_DATA{i+1}", None),
                r.get(f"EST_VAL{i+1}", None),
            ]

        # federais
        for i in range(max_fed):
            row += [
                r.get(f"FED_EMP{i+1}", ""),
                r.get(f"FED_DATA{i+1}", None),
                r.get(f"FED_VAL{i+1}", None),
            ]

        row += [r.get("TOTAL EST", 0.0), r.get("TOTAL FED", 0.0), r.get("TOTAL EMP", 0.0), r.get("Saldo", None)]
        ws.append(row)

    last_row = ws.max_row

    # freeze panes
    ws.freeze_panes = "A3"

    # styles
    thin = Side(style="thin")
    border_bottom = Border(bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    # fills (similar ao teu padrão)
    fill_header_gray = PatternFill("solid", fgColor="F2F2F2")
    fill_contract = PatternFill("solid", fgColor="E2F0D9")
    fill_est_header = PatternFill("solid", fgColor="D9E1F2")
    fill_fed_header = PatternFill("solid", fgColor="FCE4D6")
    fill_tot_header = PatternFill("solid", fgColor="E2E0F0")

    fill_data_gray = PatternFill("solid", fgColor="D9D9D9")
    fill_contract_data = PatternFill("solid", fgColor="92D050")
    fill_est_data = PatternFill("solid", fgColor="B4C6E7")
    fill_fed_data = PatternFill("solid", fgColor="FABF8F")
    fill_tot_data = PatternFill("solid", fgColor="CC99FF")
    fill_saldo_data = PatternFill("solid", fgColor="92D050")

    # header formatting
    for c in range(1, total_cols + 1):
        ws.cell(row=1, column=c).alignment = center
        ws.cell(row=2, column=c).alignment = center
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=2, column=c).font = bold

    # color header A-F
    for r in (1, 2):
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill_header_gray

    # contract header (D)
    for r in (1, 2):
        ws.cell(row=r, column=4).fill = fill_contract

    # estadual header block
    if max_est > 0:
        start = 7
        end = 7 + max_est * 3 - 1
        for r in (1, 2):
            for c in range(start, end + 1):
                ws.cell(row=r, column=c).fill = fill_est_header

    # federal header block
    if max_fed > 0:
        start = 7 + max_est * 3
        end = start + max_fed * 3 - 1
        for r in (1, 2):
            for c in range(start, end + 1):
                ws.cell(row=r, column=c).fill = fill_fed_header

    # totals header (last 4)
    tot_start = total_cols - 3
    tot_end = total_cols
    for r in (1, 2):
        for c in range(tot_start, tot_end + 1):
            ws.cell(row=r, column=c).fill = fill_tot_header

    # data colors and formats
    # A-F data gray
    for r in range(3, last_row + 1):
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill_data_gray

    # Contract col D data green
    for r in range(3, last_row + 1):
        ws.cell(row=r, column=4).fill = fill_contract_data

    # Estadual data block
    if max_est > 0:
        start = 7
        end = 7 + max_est * 3 - 1
        for r in range(3, last_row + 1):
            for c in range(start, end + 1):
                ws.cell(row=r, column=c).fill = fill_est_data

    # Federal data block
    if max_fed > 0:
        start = 7 + max_est * 3
        end = start + max_fed * 3 - 1
        for r in range(3, last_row + 1):
            for c in range(start, end + 1):
                ws.cell(row=r, column=c).fill = fill_fed_data

    # Totals data (TOTAL EST..TOTAL EMP) roxo + Saldo verde
    col_total_est = total_cols - 3
    col_total_fed = total_cols - 2
    col_total_emp = total_cols - 1
    col_saldo = total_cols

    for r in range(3, last_row + 1):
        ws.cell(row=r, column=col_total_est).fill = fill_tot_data
        ws.cell(row=r, column=col_total_fed).fill = fill_tot_data
        ws.cell(row=r, column=col_total_emp).fill = fill_tot_data
        ws.cell(row=r, column=col_saldo).fill = fill_saldo_data
   
    # -------------------------
    # formats: texto / money / date
    # -------------------------

    # CNPJ/CPF column C as text
    for r in range(3, last_row + 1):
        ws.cell(row=r, column=3).number_format = "@"

    # R$ Contrato (F)
    for r in range(3, last_row + 1):
        ws.cell(row=r, column=6).number_format = MONEY_FMT

    # Estadual: DATA e VALOR
    for i in range(max_est):
        date_col = 7 + i * 3 + 1   # EMP, DATA, VALOR -> DATA = +1
        val_col  = 7 + i * 3 + 2   # VALOR = +2
        for r in range(3, last_row + 1):
            ws.cell(row=r, column=date_col).number_format = DATE_FMT
            ws.cell(row=r, column=val_col).number_format  = MONEY_FMT

    # Federal: DATA e VALOR
    base = 7 + max_est * 3
    for i in range(max_fed):
        date_col = base + i * 3 + 1
        val_col  = base + i * 3 + 2
        for r in range(3, last_row + 1):
            ws.cell(row=r, column=date_col).number_format = DATE_FMT
            ws.cell(row=r, column=val_col).number_format  = MONEY_FMT

    # Totais + Saldo (últimas 4 colunas)
    col_total_est = total_cols - 3
    col_total_fed = total_cols - 2
    col_total_emp = total_cols - 1
    col_saldo     = total_cols

    for r in range(3, last_row + 1):
        ws.cell(row=r, column=col_total_est).number_format = MONEY_FMT
        ws.cell(row=r, column=col_total_fed).number_format = MONEY_FMT
        ws.cell(row=r, column=col_total_emp).number_format = MONEY_FMT
        ws.cell(row=r, column=col_saldo).number_format     = MONEY_FMT

    # bottom border each data row (default-ish)
    for r in range(3, last_row + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = border_bottom

    # alignment
    for r in range(1, last_row + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center")

    ## -------------------------
    # Auto width com largura mínima e máxima
    # -------------------------

    MIN_WIDTH = 20   # largura mínima
    MAX_WIDTH = 45   # largura máxima

    for c in range(1, total_cols + 1):
        max_len = 0

        # varre até 300 linhas para não pesar muito
        for r in range(1, min(last_row, 300) + 1):
            value = ws.cell(row=r, column=c).value
            if value is None:
                continue

            text = str(value)

            max_len = max(max_len, len(text))

        # largura final respeitando mínimo e máximo
        adjusted_width = max(MIN_WIDTH, min(max_len + 2, MAX_WIDTH))
        ws.column_dimensions[get_column_letter(c)].width = adjusted_width

    # export
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -----------------------------
# UI
# -----------------------------
st.set_page_config(page_title="Consolidador Empenhos SIAFIC", layout="wide")
st.title("Consolidador Empenhos SIAFIC → Relatório Consolidado")

st.write("1) Informe o **Número de Protocolo / Número do GMS / Número do Contrato / Valor do Contrato**.")
st.write("2) Anexe o **relatório do SIAFIC**.")
st.write("3) Baixe o consolidado.")
st.write("4) Seja Feliz :)")

st.subheader("Entrada: Protocolos")
default_input = pd.DataFrame(
    [
        {"Protocolo": "", "GMS": "", "Contrato": "", "Valor do Contrato": ""},
    ]
)
prot_df = st.data_editor(
    default_input,
    num_rows="dynamic",
    use_container_width=True
)

st.subheader("Upload: Relatório SIAFIC (XLSX ou CSV)")
uploaded = st.file_uploader("Selecione o arquivo", type=["xlsx", "xls", "csv"])

st.caption("Feito por Eduardo Portella")

if uploaded:
    try:
        name = uploaded.name.lower()

        # -------------------------
        # 1) Leitura CRUA (sem header)
        # -------------------------
        df_raw = None

        if name.endswith(".csv"):
            df_raw = pd.read_csv(uploaded, header=None, dtype=object, low_memory=False)

        elif name.endswith(".xls"):
            # .xls (binário) -> xlrd
            df_raw = pd.read_excel(uploaded, header=None, dtype=object, engine="xlrd")

        else:
            # tenta .xlsx normal
            try:
                df_raw = pd.read_excel(uploaded, header=None, dtype=object, engine="openpyxl")
            except ValueError as e:
                # caso clássico: "0 worksheets found"
                if "0 worksheets found" in str(e).lower():
                    df_raw = read_weird_ooxml_xlsx(uploaded)
                else:
                    raise

        # -------------------------
        # 2) Limpeza (A4 + corta TOTAL/rodapé)
        # -------------------------
        df_siafic = clean_siafic_table(df_raw, start_row=3)

        if df_siafic is None or df_siafic.empty:
            st.error("Não foi possível identificar a tabela no arquivo (após remover cabeçalho e rodapé).")
            st.stop()

        # -------------------------
        # 3) Auto-detect colunas
        # -------------------------
        mapping = guess_columns(df_siafic)
        ok, missing = validate_mapping(mapping, df_siafic)

        # -------------------------
        # 4) Geração do consolidado
        # -------------------------
        if ok and st.button("Gerar planilha consolidada", type="primary"):
            prot_df2 = prot_df.copy()

            if "Valor do Contrato" in prot_df2.columns:
                prot_df2 = prot_df2.rename(columns={"Valor do Contrato": "Valor do Contrato"})

            if "Valor do Contrato" not in prot_df2.columns:
                prot_df2["Valor do Contrato"] = ""

            df_out, max_est, max_fed = build_consolidated(
                df_siafic=df_siafic,
                protocolos_input=prot_df2,
                mapping=mapping
            )

            money_cols = ["R$ Contrato", "TOTAL EST", "TOTAL FED", "TOTAL EMP", "Saldo"]

            # adiciona também todas as colunas EST_VALx e FED_VALx existentes
            money_cols += [c for c in df_out.columns if c.startswith("EST_VAL") or c.startswith("FED_VAL")]

            for c in money_cols:
               if c in df_out.columns:
                  df_out[c] = pd.to_numeric(df_out[c], errors="coerce")

            xlsx_bytes = write_excel_formatted(df_out, max_est, max_fed)

            now = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.success("Consolidado gerado com sucesso.")
            st.download_button(
                "Baixar XLSX consolidado",
                data=xlsx_bytes,
                file_name=f"REL_CONSOLIDADO_{now}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.subheader("Prévia (primeiras linhas)")
            st.dataframe(df_out.head(50), use_container_width=True)

    except Exception as e:
        st.exception(e)



